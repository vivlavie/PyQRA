#PyLeakRateFreq1.py
#Calculate Leak Frequency that give leak rate within a certain range in specific module

from openpyxl import load_workbook
import math
import numpy as np
import matplotlib.pyplot as plt
from scipy.interpolate import interp1d
import scipy.ndimage
import dill
import sys
import matplotlib.pylab as pltparam


def print_a_event(lEvent,key='',hole='',weather=''):
    i=0
    for e in lEvent:
        if (key in e.Key) and (hole in e.Hole) and (weather in e.Weather):
            print(i,e.Key)
        i += 1


element_dump_filename = 'Bv06_dump'
Area = "ProcessArea"

# element_dump_filename = 'Bv06_offloading_dump'
# Area = "Offloading"


with open(element_dump_filename,'rb') as element_dump:
    lEvent = dill.load(element_dump)
    Loaded = True



iIS=load_workbook(filename='IS_v12_shk.xlsx')
shIS = iIS['Isolatable summary']
IS_sub = {}
numESDVs = {}
Modules = {}
Deck = {}
r = 3
# while r < 79:
#     nsub = shIS.cell(r,4).value
#     IS_sub[shIS.cell(r,3).value] = [r,nsub]
#     r += nsub
while r < 82:
    pv = shIS.cell(r,5).value
    IS_sub[pv] = shIS.cell(r,11).value #Read for each leak at respective height
    Modules[pv] = shIS.cell(r,7).value
    Deck[pv] =  shIS.cell(r,8).value
    if shIS.cell(r,24).value != None:
       nedvs = shIS.cell(r,24).value.count("\"")    
       numESDVs[pv] = nedvs
    else:
       numESDVs[pv] = pnedvs
    pnedvs = nedvs
    r += 1


# Targets = ['UtilityModule','LQ','SMA']

Targets = ['GTG_CP2','GTG_Int2','GTG_Exh2','LQ_Intake','EnI_Intake','Wks_Intake','EnI_CDr','EnI_PDr','Wks_CDr','Wks_SDr','Helideck','LB_S','LB_P','SMA_Emb','AFT_CC','FWD_CC']

Wt0=[[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
[0,0,1,0,0,0,0,0,0,0,0,0,0,0,5,0],
[1,1,8,0,2,0,0,0,0,0,1,0,0,0,11,0]]


Conditions = {}
NumCases = 24
Conditions['Leak 1 - Small'] = [[1,40], ['S02','P02','S03','P03'], [0, 1, 0],'-G']
Conditions['Leak 1 - Medium'] = [[40,105], ['S02','P02','S03','P03'], [8, 3, 0],'-G']
Conditions['Leak 1 - Large'] = [[105,1000], ['S02','P02','S03','P03'], [11, 11, 0],'-G']

Conditions['Leak 2 - Small'] = [[1,10], ['S05','P05'], [0, 0, 2],'-G']
Conditions['Leak 2 - Medium'] = [[10,40], ['S05','P05'], [0, 0, 4],'-G']
Conditions['Leak 2 - Large'] = [[40,1000], ['S05','P05'], [3, 5, 4],'-G']

Conditions['Leak 3 and 4 - Small'] = [[1,10], ['S04','P04'], [0, 0, 0],'-G']
Conditions['Leak 3 and 4 - Medium'] = [[10,40], ['S04','P04'], [5, 4, 4],'-G']
Conditions['Leak 3 and 4 - Large'] = [[40,1000], ['S04','P04'], [12, 15, 4],'-G']

Conditions['Leak 5 - Small'] = [[1,10], ['S04','P04','S05','P05'], [0, 0, 0],'-L']
Conditions['Leak 5 - Medium'] = [[10,40], ['S04','P04','S05','P05'], [1, 2, 4],'-L']
Conditions['Leak 5 - Large'] = [[40,1000], ['S04','P04','S05','P05'], [15, 11, 3],'-L']


# LeakRateThreholds = [40, 105, 105, 105]
# Modules = [['S03','P03'], ['S03','P03'], ['S04','P04'], ['S05','P05']]
# FlamGasImpWeight = [[1/6, 0, 0], [1, 1/6, 0],[0, 0, 0], [0, 0, 0]]

Fsum = np.array([0.,0.,0.])
for k in Conditions:
    c = Conditions[k]
    lr_low = c[0][0]
    lr_high = c[0][1]
    modules = c[1]
    weight = c[2]
    gas_or_liquid = c[3]
    F = np.array([0.,0.,0.])
    # F=[0,0,0]
    for e in lEvent:
        # if not ("-L" in e.Key): #only for light gas
        if (gas_or_liquid in e.Key): #only for light gas
            # print(e.Key)
            pv,hole,weather = e.Key.split('\\')    
            mod = e.Module
            elf = e.Frequency
            if gas_or_liquid == "-L":
                elr = e.Discharge.ReleaseRate*max(0.05,1-e.Discharge.LiquidMassFraction)
                # elr = e.Discharge.ReleaseRate
            else:
                elr = e.Discharge.ReleaseRate
            if (mod in modules) and ((elr >= lr_low) and (elr < lr_high)):
                for t in range(len(Targets)):
                    w = c[2][t]/NumCases
                    F[t] += w*elf
    print("{:40s} [{:8.2e}, {:8.2e}, {:8.2e}]".format(k,F[0],F[1],F[2]))
    Fsum = Fsum + F
print("{:40s} [{:8.2e}, {:8.2e}, {:8.2e}]".format("Total",Fsum[0],Fsum[1],Fsum[2]))



rrmax = 0
for e in lEvent:
    if not ("-L" in e.Key) and (e.Discharge.ReleaseRate > rrmax):
        rrmax = e.Discharge.ReleaseRate
        print(e.Module,rrmax)
print(rrmax)

for e in lEvent:
    if (not ("-G" in e.Key)) and (not ("-L" in e.Key)):        
        print(e.Key, e.Module)
