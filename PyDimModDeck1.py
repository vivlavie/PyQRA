#PyDimModeDeck
#Dim jet fire length at 5 minues after the release

from openpyxl import load_workbook
import math
import numpy as np
import matplotlib.pyplot as plt
from scipy.interpolate import interp1d
import scipy.ndimage
from scipy.ndimage.filters import gaussian_filter
import dill
import sys
import matplotlib.pylab as pltparam

def jffit(m):
    jl_lowe = 2.8893*np.power(55.5*m,0.3728)
    # if m>5:
    #     jf = -13.2+54.3*math.log10(m)
    # elif m>0.1:
    #     jf= 3.736*m + 6.
    # else:
    #     jf = 0.
    # print(m, jl_lowe,jf)
    return jl_lowe
def mfit(jl):
    m = np.power(10,math.log10(jl/2.8893)/0.3728) / 55.5
    return m

def print_cum_cube(cube,AA):
    F=0.
    print("{:35s} Mod  (Freq. ) - Jet length  CumFreq".format(cube))
    for e in AA[::-1]:
        # pv,hole,weather = e[2].split("\\")
        # hole = hole.split("_")[0]
        F += e[1]
        pv,hole,weather = e[2].split("\\")
        mod = Modules[pv]
        print("{:35s}{:5s} {:8.2e} {:8.1f}   {:8.2e}".format(e[2],mod,e[1],e[0],F))
        if F>1.0E-3:
            break
def print_cum_cube_file(cube,AA,a_file):
    F=0.
    print("{:35s} Mod  (Freq. ) - Jet length  CumFreq".format(cube),file = a_file)
    for e in AA[::-1]:
        # pv,hole,weather = e[2].split("\\")
        # hole = hole.split("_")[0]
        F += e[1]
        pv,hole,weather = e[2].split("\\")
        mod = Modules[pv]
        print("{:35s}{:5s} {:8.2e} {:8.1f}   {:8.2e}".format(e[2],mod,e[1],e[0],F),file = a_file)
        if (F>1.0E-3) or (e[0] == 0.):
            break
def print_a_event(lEvent,key='',hole='',weather=''):
    i=0
    for e in lEvent:
        if (key in e.Key) and (hole in e.Hole) and (weather in e.Weather):
            print(i,e.Key)
        i += 1
        

WantPlot = True
cube_result2file = True


element_dump_filename = 'Bv06_dump'
Area = "ProcessArea"

# element_dump_filename = 'Bv06_offloading_dump'
# Area = "Offloading"


if Loaded == False:
    with open(element_dump_filename,'rb') as element_dump:
        lEvent = dill.load(element_dump)
        Loaded = True



iIS=load_workbook(filename='IS_v12_shk.xlsx')
shIS = iIS['Isolatable summary']
IS_sub = {}
numESDVs = {}
Modules = {}
Deck = {}
r = 3
# while r < 79:
#     nsub = shIS.cell(r,4).value
#     IS_sub[shIS.cell(r,3).value] = [r,nsub]
#     r += nsub
while r < 82:
    pv = shIS.cell(r,5).value
    IS_sub[pv] = shIS.cell(r,11).value #Read for each leak at respective height
    Modules[pv] = shIS.cell(r,7).value
    Deck[pv] =  shIS.cell(r,8).value
    if shIS.cell(r,24).value != None:
       nedvs = shIS.cell(r,24).value.count("\"")    
       numESDVs[pv] = nedvs
    else:
       numESDVs[pv] = pnedvs
    pnedvs = nedvs
    r += 1


# Jet fire length analysis for each module
numDirections = 6
Ts = [0, 300, 600, 1800, 3600]
# for ti in range(0,5): #tt 1: 5 min, 2:10 min, 3: 30 min, 3: 60min
for ti in range(1,5):
    print(Ts[ti])
    F = 0.
    JFModule = {}
    JFLModule = {}
    IDAsorted = []
    for e in lEvent:
        pv,hole,weather = e.Key.split('\\')
        mod = e.Module+'_'+e.Deck
        # mod = e.Module
        jfl = e.JetFire.JetLengths[ti]
        jff = e.JetFire.Frequency*numDirections #considering it may happn in any direction
        # jff = e.JetFire.Frequency #considering directionality
        F += jff
        if mod in JFLModule.keys():
            JFLModule[mod].append([jfl,jff,e.Key])              
        else:
            JFLModule[mod] = [[jfl,jff,e.Key]]

    print("{:20s} Toal jet fire frequency: {:10.2e}".format(Area, F))
    print("{:20s}{:30s}{:10s}{:20s}".format("Module","Scenario","Jet length ","Rel rate[kg/s]"))

    # for m in JFLModule.keys():
    for m in ['P05_A']:
        IDAsorted = sorted(JFLModule[m], key = lambda fl: fl[0]) #with the longest duration at the bottom
        
        fn = "DimScnsModule\\{}_{:4d}".format(m,Ts[ti])
        if cube_result2file == True:
            f_cube_result = open(fn+".txt","w")
        
        cf = 0
        jp = IDAsorted[-1]
        DimFreq = 1.0E-4
        di = 0
        scn = ''
        cfp = jp[1]#frequency for the longest duration jet fire
        InterpolationSuccess = False
        if cfp > DimFreq:
            di = jp[0]
            scn = jp[2]
        else:
            for j in IDAsorted[-2:0:-1]:
                cf = cfp + j[1]
                dn = j[0]
                scn = j[2]
                if (cf >= DimFreq) and (cfp < DimFreq) and (dn > 0):
                    dp = jp[0]
                    dn = j[0]
                    # di = (dn-dp)/(cf-cfp)*(DimFreq - cfp) + dp
                    # print('Dimensioning jet duration {:8.1f}'.format(di))
                    scnp = jp[2]
                    scn = j[2]
                    InterpolationSuccess = True
                    #Choose the scenario close to the threshold
                    if abs(cf-DimFreq) > abs(cfp-DimFreq):
                        scn = scnp
                        di = dp
                    else:
                        di = dn
                    break
                cfp = cf
                # print(cfp,cf)
                jp = j
        if di == 0:
            rr = 0
            scn = ''
        else:
            for e in lEvent:
                if scn == e.Key:
                    rr = mfit(di/e.jfscale)
                    # print(e.JetFire.JetLengths)
            print("{:20s}{:30s}{:10.1f}{:20.1f}".format(m,scn,di,rr))
        # print_cum_cube(mod,IDAsorted)
        
        
        if (cube_result2file == True):
            if InterpolationSuccess:                
                print_cum_cube_file(fn,IDAsorted,f_cube_result) 
            else:                
                print_cum_cube_file(fn+" No dimensioning scenario ",IDAsorted,f_cube_result)        
        

        if (WantPlot == True):
            ll = len(IDAsorted)
            ec = np.zeros([ll,2])
            i=0
            ec[i,1] = IDAsorted[-1][0] #the longest duration
            ec[i,0] = IDAsorted[-1][1] #Frequency for the longest duration
            for i in range(1,len(IDAsorted)):
                ec[i,1] = IDAsorted[ll-i-1][0]
                ec[i,0] = ec[i-1,0] + IDAsorted[ll-i-1][1]

            # plt.figure(figsize=(5.91, 3.15))
            CF = ec[1:,0]
            JFL = ec[1:,1]
            masscolor = 'tab:blue'
            fig,ax1 = plt.subplots()
            ax1.set_xlabel('Jet Fire Length [m]')
            ax1.set_ylabel('Cumulative Frequency [#/year]',color=masscolor)
            ax1.semilogy(JFL,CF,color=masscolor)
            ax1.set_ylim(top=5E-4,bottom=1E-6)
            ax1.set_xlim(left=0)
            # ax1.set_xlim(left=0, right=3600)
            ax1.tick_params(axis='y',labelcolor=masscolor)
            # ax1.xaxis.set_major_locator(plt.FixedLocator([300, 600, 1800, 3600]))
            ax1.annotate(scn,xy=(di,1.0E-4),xytext=(di,2E-4),horizontalalignment='left',verticalalignment='top',arrowprops = dict(facecolor='black',headwidth=4,width=2,headlength=4))
            # ax.xaxis.set_major_formatter(plt.FixedFormatter(['2/3','3/4','4/5','S05']))
            # ax.yaxis.set_major_locator(plt.FixedLocator([-27, -3.1, 3.1, 27]))
            # ax.yaxis.set_major_formatter(plt.FixedFormatter(['ER_S','Tray_S','Tray_P','ER_P']))
            ax1.grid(True,which="major")

            if InterpolationSuccess:                
                plt.title("{:4s} at {:4.0f} sec - Jet fire from {:20s} of {:6.1f} [m]".format(m,Ts[ti],scn,di))
            else:
                plt.title("{:4s} at {:4.0f} sec - No dimensioning fire".format(m,Ts[ti]))            
            plt.tight_layout()
            plt.show()

            fig.savefig(fn+".png")
            plt.close()


        if cube_result2file == True:
            f_cube_result.close()


