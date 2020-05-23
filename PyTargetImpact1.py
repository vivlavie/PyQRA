#PyTargetImpact1.py/
#To assess how often a target be affected by a certian criterion

#PyDimCube4 using 'Bv06_c.xlsx'

#Deluge skids
DS = {}
DS['071-SZ-001'] =[	161.091,	-20.869,	29.325]
DS['071-SZ-002'] =[	82.545,	-20.819,	29.345]
DS['071-SZ-003'] =[	91.472,	-20.669,	29.319]
DS['071-SZ-004'] =[	178.207,	-20.394,	29.353]
DS['071-SZ-005'] =[	76.348,	4.144,	29.625]
DS['071-SZ-006'] =[	76.826,	-10.144,	29.605]
DS['071-SZ-007'] =[	127.409,	20.032,	29.341]
DS['071-SZ-008'] =[	185.860,	-20.434,	29.320]
#The following are actually cargo tank manifold

DS['Piping in CT6 (P) Area'] =[	68.89,	9.94,	29.71]
DS['Piping in CT6 (S) Area'] =[	70.47,	-10.75,	29.71]
DS['Piping in CT5 (P) Area'] =[	84.27,	8.29,	29.71]
DS['Piping in CT5 (S) Area'] =[	84.27,	-10.63,	29.71]
DS['Piping in CT4 (P) Area'] =[	110.38,	8.51,	29.71]
DS['Piping in CT4 (S) Area'] =[	110.38,	-10.58,	29.71]
DS['Piping in CT3 (P) Area'] =[	136.32,	8.31,	29.71]
DS['Piping in CT3 (S) Area'] =[	136.32,	-10.6,	29.71]
DS['Piping in CT2 (P) Area'] =[	158.06,	8.326,	29.71]
DS['Piping in CT2 (S) Area'] =[	158.06,	-10.46,	29.71]
DS['Piping in CT1 (P) Area'] =[	180.01,	8.135,	29.71]
DS['Piping in CT1 (S) Area'] =[	180.01,	-10.5,	29.71]


from openpyxl import load_workbook
import math
import numpy as np
import matplotlib.pyplot as plt
from scipy.interpolate import interp1d
import scipy.ndimage
from scipy.ndimage.filters import gaussian_filter
import dill
import sys

import matplotlib.pylab as pltparam


# sys.stdout = open('DimCube2.txt','w')

WantPlot = False
cube_result2file = False
FirewallX = 140.7 #meter from AP

def DtoH (D05):
    if D05 > 19.7:
            H05 = 0.4664*D05 + 18.345
    elif D05 < 1.0:
        H05 = 0.
    else: #  1.0 < D05 < 19.7:
        H05 = 1.2592*D05 + 2.7235
    return H05
def PD(t,De,Te):
    if t < Te:
        PD = De * (1-math.sqrt(t/Te))
    else:
        PD = 0
    return PD
#pool burning rate 0.062 kg/m2-s
#MS ; spilt mass
#Pool fire duration
# Ms / (3.14*D^2/4 * br) *3/2 where 3/2 is a factor to consider shringking pool

def print_cum_cube(cube,AA):
    F=0.
    print("{:35s} Mod  (Freq. ) - Jet Duration  CumFreq".format(cube))
    for e in AA[::-1]:
        # pv,hole,weather = e[2].split("\\")
        # hole = hole.split("_")[0]
        F += e[1]
        pv,hole,weather = e[2].split("\\")
        mod = Modules[pv]
        print("{:35s}{:5s} {:8.2e} {:8.1f}   {:8.2e}".format(e[2],mod,e[1],e[0],F))
        if F>1.0E-3:
            break
def print_cum_cube_file(cube,AA,a_file):
    F=0.
    print("{:35s} Mod  (Freq. ) - Jet Duration  CumFreq".format(cube),file = a_file)
    for e in AA[::-1]:
        # pv,hole,weather = e[2].split("\\")
        # hole = hole.split("_")[0]
        F += e[1]
        pv,hole,weather = e[2].split("\\")
        mod = Modules[pv]
        print("{:35s}{:5s} {:8.2e} {:8.1f}   {:8.2e}".format(e[2],mod,e[1],e[0],F),file = a_file)
        if F>1.0E-3:
            break


# element_dump_filename = 'Bv06_dump'
# icubeloc='SCE_CUBE_XYZ2_Process'

element_dump_filename = 'Bv06_hull_dump'
icubeloc='SCE_CUBE_XYZ2_HullDeck'

# element_dump_filename = 'Bv06_offloading_dump'
# icubeloc='SCE_CUBE_XYZ2_Offloading'

# element_dump_filename = 'Bv06_utility_dump'
# icubeloc='SCE_CUBE_XYZ2_Utility'

with open(element_dump_filename,'rb') as element_dump:
    lEvent = dill.load(element_dump)

#Read 'Presure vessel' from Input
iExlFilename='Bv06_i'
iExl=load_workbook(filename=iExlFilename+'.xlsx')
shPV = iExl['Pressure vessel']
X = {}
Y = {}
Z = {}
r=63
while shPV.cell(r,1).value == "Yes":    
    study  = shPV.cell(r,2).value
    pv  = shPV.cell(r,8).value    
    # key = study  + "\\" +  pv
    key = pv    
    X[key] = shPV.cell(r,136).value
    Y[key] = shPV.cell(r,137).value    
    Z[key] = shPV.cell(r,20).value        #Elevation
    r += 1
numPV = r-63
def jffit(m):
    jl_lowe = 2.8893*np.power(55.5*m,0.3728)
    if m>5:
        jf = -13.2+54.3*math.log10(m)
    elif m>0.1:
        jf= 3.736*m + 6.
    else:
        jf = 0.
    # print(m, jl_lowe,jf)
    return jl_lowe


CubeImpingeDuration = {}
Cubes = list(DS.keys())
Xcube = {}
Ycube = {}
Zcube = {}
for k in DS.keys():    
    Xcube[k] = DS[k][0]
    Ycube[k] = DS[k][1]
    Zcube[k] = DS[k][2]
    

print("{:32s} {:5s} {:4s} {:10s} Heat[kW/m2]  ".\
                format('Scenario','Module','Deck','Deluge Skid'))
ImpingeDurationArray = []
HeatCriterion1 = 4.7
HeatCriterion2 = 12.5
HeatCriterion3 = 37.5
# for id in [list(DS.keys())[0]]:    
for id in DS.keys():    
# for i in range(0,37):
# for i in [ncube-1]:
    # li = 0
    # sep = np.zeros((len(lEvent),2))
    ImpingeDurationArray = []    
    xx = Xcube[id]
    yy = Ycube[id]
    zz = Zcube[id]
    if cube_result2file == True:
        f_cube_result = open(id+".txt","w")
    # if (zz >= 35.0) and (zz < 44):
    #     CubeDeck = 'A'
    # elif (zz >= 44.0) and (zz < 53):
    #     CubeDeck = 'B'
    # elif (zz >= 53):
    #     CubeDeck = 'C'
    # elif (zz < 35 ):
    #     CubeDeck = 'Hull'
    
    for e in lEvent:        
        
        #Exposure to jet fire
        #read frequency
        # sep[li,1] = e.JetFire.Frequency
        #check if the receptor is in the flame
        dx = xx - e.X
        dy = yy - e.Y
        dz = zz - (e.ReleaseHeight+35)
        rr = math.sqrt(dx*dx+dy*dy+dz*dz)
        #Jet fire
        d04 = e.JetFire.D04
        d12 = e.JetFire.D12            
        if rr < e.JetFire.Length:
            q = e.JetFire.SEP        
        elif (d12 != 'n/a') and (rr < d12 ):
        # elif (d12 != 'n/a') and (d04 != 'n/a'):
            q = (e.JetFire.SEP-12.5)/(e.JetFire.Length-d12)*(rr-d12)+12.5
        elif (d04 != 'n/a') and (rr < d04 ):
            q = (4-12.5)/(d04-d12)*(rr-d12)+12.5            
        elif (d04 != 'n/a') and (d12 != 'n/a') and (rr >= d04) :
            q = max((4-12.5)/(d04-d12)*(rr-d12)+12.5,0)
        else:
            q = 0
        if (q<0):
            print(e.Key, "something wrong: q = ",q)
        ImpingeDurationArray.append([q,e.JetFire.Frequency,e.Key+"_Jet"])
        #Early pool fire
        if e.EarlyPoolFire != None:
            d04 = e.EarlyPoolFire.D04
            d12 = e.EarlyPoolFire.D12            
            if rr < e.EarlyPoolFire.Diameter:
                q = e.EarlyPoolFire.SEP
            elif (d12 != 'n/a') and (rr < d12 ):
            # elif (d12 != 'n/a') and (d04 != 'n/a'):
                q = (e.EarlyPoolFire.SEP-12.5)/(e.EarlyPoolFire.Diameter-d12)*(rr-d12)+12.5
            elif (d04 != 'n/a') and (rr < d04 ):
                q = (4-12.5)/(d04-d12)*(rr-d12)+12.5            
            elif (d04 != 'n/a') and (d12 != 'n/a') and (rr >= d04) :
                q = max((4-12.5)/(d04-d12)*(rr-d12)+12.5,0)
            else:
                q = 0
            if (q<0):
                print(e.Key, "something wrong: q = ",q)
            
            ImpingeDurationArray.append([q,e.EarlyPoolFire.Frequency,e.Key+"_EP"])
        #Late pool fire
        if e.LatePoolFire != None:
            d04 = e.LatePoolFire.D04
            d12 = e.LatePoolFire.D12            
            if rr < e.LatePoolFire.Diameter:
                q = e.LatePoolFire.SEP
            elif (d12 != 'n/a') and (rr < d12 ):
            # elif (d12 != 'n/a') and (d04 != 'n/a'):
                q = (e.LatePoolFire.SEP-12.5)/(e.LatePoolFire.Diameter-d12)*(rr-d12)+12.5
            elif (d04 != 'n/a') and (rr < d04 ):
                q = (4-12.5)/(d04-d12)*(rr-d12)+12.5            
            elif (d04 != 'n/a') and (d12 != 'n/a') and (rr >= d04) :
                q = max((4-12.5)/(d04-d12)*(rr-d12)+12.5,0)
            else:
                q = 0
            if (q<0):
                print(e.Key, "something wrong: q = ",q)
            
            ImpingeDurationArray.append([q,e.LatePoolFire.Frequency,e.Key+"_LP"])
                                    
      
    #To pin-point a scenario that give the dimensioning scenario
    IDAsorted = sorted(ImpingeDurationArray, key = lambda fl: fl[0]) #with the longest duration at the bottom
    
    cf = 0
    jp = IDAsorted[-1]
    DimFreq = 1.0E-4
    di = 0
    cfp = jp[1]#frequency for the longest duration jet fire
    InterpolationSuccess = False
    HeatCriterionCheck1 = False
    HeatCriterionCheck2 = False
    HeatCriterionCheck3 = False
    cfp1,cfp2,cfp3 = 0.,0.,0.
    if cfp > DimFreq:
        di = jp[0]
        scn = jp[2]
    else:
        for j in IDAsorted[-2:0:-1]:
            cf = cfp + j[1]
            dp = jp[0]
            dn = j[0]
            scn = j[2]
            if cf >= DimFreq and cfp < DimFreq:
                # di = (dn-dp)/(cf-cfp)*(DimFreq - cfp) + dp
                # print('Dimensioning jet duration {:8.1f}'.format(di))
                scnp = jp[2]
                scn = j[2]
                InterpolationSuccess = True
                #Choose the scenario close to the threshold
                if abs(cf-DimFreq) > abs(cfp-DimFreq):
                    scn = scnp
                    di = dp
                else:
                    di = dn
                break
            cfp = cf
                        
            #For each accident, check if the Heatcriterion is exceeded. Heat is in the descending order. Criterion3 is met at first.
            if (dn < HeatCriterion1) and (HeatCriterionCheck1 == False):
                # print("Cumulative frequency for Heat {:8.1f} kW/m2 exceeded at {:s} is {:6.1e}".format(HeatCriterion1,scn,cfp))
                cfp1 = cfp
                HeatCriterionCheck1 = True
            if (dn < HeatCriterion2) and (HeatCriterionCheck2 == False):
                # print("Cumulative frequency for Heat {:8.1f} kW/m2 exceeded at {:s} is {:6.1e}".format(HeatCriterion2,scn,cfp))
                cfp2 = cfp                
                HeatCriterionCheck2 = True
            if (dn < HeatCriterion3) and (HeatCriterionCheck3 == False):
                # print("Cumulative frequency for Heat {:8.1f} kW/m2 exceeded at {:s} is {:6.1e}".format(HeatCriterion3,scn,cfp))
                cfp3 = cfp
                HeatCriterionCheck3 = True
            # print(cfp,cf)
            #The following print() will give an error if either of cfp1,cfp2 or cfp3 is not evaluated.
            jp = j
        print("{:40s}{:10.1e}{:10.1e}{:10.1e}".format(id,cfp3,cfp2,cfp1))
    
    # print(scn)
    if InterpolationSuccess:
        CubeImpingeDuration[Cubes[i]] = di        
        pv,hole,weather = scn.split("\\")            
        weather_fire = weather            
        x1 = X[pv]
        y1 = Y[pv]
        z1 = Z[pv]
        for e in lEvent:
            if e.Key in scn:
                break
        rrifound = False
        rri = 0
        rr5min = 0.
        for i in range(0,len(e.TVD)-1):
            if (e.TVD[i,0] < di) and (e.TVD[i+1,0] >= di):
                rri = e.TVD[i,2]
                rr5min = e.Discharge.RRs[1]
                rrifound = True
                break
        if rrifound == False:
            print(id,"Something wrong, rri not found", di, e.TVD[-1,0])

        z1 = z1+35
        dx=xx-x1
        dy=yy-y1
        dz=zz-z1
        jli = jffit(rri)
        xm = x1 + 0.61*dx
        ym = y1 + 0.61*dy
        zm = z1 + 0.61*dz
        # ll = math.sqrt(dx*dx+dy*dy+dz*dz)
        # print("IS: {:35s} {:4s}{:2s}{:6.1f}{:6.1f}{:6.1f}  Cube: {:10s}{:6.1f}{:6.1f}{:6.1f} Dx:{:6.1f}{:6.1f}{:6.1f} {:8.1f}{:8.1f}".\
        #     format(scn,Modules[pv],Deck[pv],x1,y1,z1,id,xx,yy,zz,xx-x1,yy-y1,zz-z1,di,rri))
        print("{:35s} {:5s} {:4s} {:10s} {:8.1f} {:8.1f} {:8.1f} {:8.1f}".\
            format(scn, Modules[pv],Deck[pv],id,di,di/60,rri,rr5min))
        if cube_result2file == True:                
            print_cum_cube_file(id,IDAsorted,f_cube_result) 
          
    elif not InterpolationSuccess:
        print("No dimensioning scenario",id)
        scn = 'No dimensioning scenario for ' + id
        # print_cum_cube(id+" No dimensioning scenario ",IDAsorted)        
        if cube_result2file == True:
            print_cum_cube_file(id+" No dimensioning scenario ",IDAsorted,f_cube_result)        


    # if (InterpolationSuccess == True) and (WantPlot == True):
    if (WantPlot == True):
        ll = len(IDAsorted)
        ec = np.zeros([ll,2])
        i=0
        ec[i,1] = IDAsorted[-1][0] #the longest duration
        ec[i,0] = IDAsorted[-1][1] #Frequency for the longest duration
        for i in range(1,len(IDAsorted)):
            ec[i,1] = IDAsorted[ll-i-1][0]
            ec[i,0] = ec[i-1,0] + IDAsorted[ll-i-1][1]

        # plt.figure(figsize=(5.91, 3.15))
        CF = ec[1:,0]
        JFL = ec[1:,1]
        masscolor = 'tab:blue'
        fig,ax1 = plt.subplots()
        ax1.set_xlabel('Heat [kW/m2]')
        ax1.set_ylabel('Cumulative Frequency [#/year]',color=masscolor)
        ax1.semilogy(JFL,CF,color=masscolor)
        ax1.set_ylim(top=5E-4,bottom=1E-6)
        ax1.set_xlim(left=0, right=3600)
        ax1.tick_params(axis='y',labelcolor=masscolor)
        ax1.xaxis.set_major_locator(plt.FixedLocator([300, 600, 1800, 3600]))
        ax1.annotate(scn,xy=(di,1.0E-4),xytext=(di,2E-4),horizontalalignment='left',verticalalignment='top',arrowprops = dict(facecolor='black',headwidth=4,width=2,headlength=4))
        # ax.xaxis.set_major_formatter(plt.FixedFormatter(['2/3','3/4','4/5','S05']))
        # ax.yaxis.set_major_locator(plt.FixedLocator([-27, -3.1, 3.1, 27]))
        # ax.yaxis.set_major_formatter(plt.FixedFormatter(['ER_S','Tray_S','Tray_P','ER_P']))
        ax1.grid(True,which="major")

        if InterpolationSuccess:
            plt.title("Cube {:10s} - {:5s} fire from {:30} for {:8.1f} [sec]".format(id,weather_fire[6:],pv+"_"+hole+"_"+weather_fire[:5],di))            
        else:
            plt.title("Cube {:10s} - No dimensioning fire".format(id))            
        plt.tight_layout()
        plt.show()

        fig.savefig("{}.png".format(id))
        plt.close()

    if cube_result2file == True:
        f_cube_result.close()
# print_cum_cube(IDAsorted)

# for c in Cubes:
#     print("{:20s} {:8.1f} [min]".format(c,CubeImpingeDuration[c]/60))
# sys.stdout.close()