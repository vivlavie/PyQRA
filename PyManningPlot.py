from openpyxl import load_workbook
import math
import numpy as np
import matplotlib.pyplot as plt
from scipy.interpolate import interp1d
import scipy.ndimage
from scipy.ndimage.filters import gaussian_filter
import dill

import matplotlib.pylab as pltparam

pltparam.rcParams["figure.figsize"] = (8,3)
pltparam.rcParams['lines.linewidth'] = 2
# pltparam.rcParams['lines.color'] = 'r'
pltparam.rcParams['axes.grid'] = True 


iExlFilename='ruby_manning'
iExl=load_workbook(filename=iExlFilename+'.xlsx')

Cases = ['Base Case','Max Case','Transit Case','Sensitivity Case']
Areas = ['S00','P01','S01','P02','S02','P03','S04','P04','S04','P05','S05','LD1','LD2','Turret','OffloadingArea','KOD','CrowNest','ER','FWDPR','CCR','Cabin','RecreationRoom','MessRoom']
Groups = [["LQ_Officers",1,14], ["CCR_Bridge_Operators",15,17],["Process_Utility_Operators",18,30],["Marine_Operators",31,35]]
Coords = {}
Coords['S00'] = [25, -25, 35, 20, 25, 35] #x,t,z,dx,dy,dz [m]
Coords['S01'] = [45, -25, 35, 20, 25, 35] #x,t,z,dx,dy,dz [m]
Coords['S02'] = [65, -25, 35, 30, 25, 35] #x,t,z,dx,dy,dz [m]
Coords['S03'] = [95, -25, 35, 40, 25, 35] #x,t,z,dx,dy,dz [m]
Coords['S04'] = [135, -25, 35, 30, 25, 35] #x,t,z,dx,dy,dz [m]
Coords['S05'] = [165, -25, 35, 30, 25, 35] #x,t,z,dx,dy,dz [m]
Coords['P00'] = [25, 0, 35, 20, 25, 35] #x,t,z,dx,dy,dz [m]
Coords['P01'] = [45, 0, 35, 20, 25, 35] #x,t,z,dx,dy,dz [m]
Coords['P02'] = [65, 0, 35, 30, 25, 35] #x,t,z,dx,dy,dz [m]
Coords['P03'] = [95, 0, 35, 40, 25, 35] #x,t,z,dx,dy,dz [m]
Coords['P04'] = [135, 0, 35, 30, 25, 35] #x,t,z,dx,dy,dz [m]
Coords['P05'] = [165, 0, 35, 30, 25, 35] #x,t,z,dx,dy,dz [m]
Coords['LD1'] = [45, 20, 35, 10, 55, 2] #x,t,z,dx,dy,dz [m]
Coords['LD2'] = [195, -25, 35, 10, 5, 2] #x,t,z,dx,dy,dz [m]
Coords['Turret']= [195, -25, 0, 20, 50, 19] #x,t,z,dx,dy,dz [m]
Coords['OffloadingArea'] = [-5, -25, 35, 15, 50, 2] #x,t,z,dx,dy,dz [m]
Coords['KOD'] = [195, 15, 35, 10, 10, 55] #x,t,z,dx,dy,dz [m]
Coords['CrowNest'] = [230, -2, 37, 5, 4, 3] #x,t,z,dx,dy,dz [m]
Coords['ER'] = [10, -25, 0, 30, 50, 19] #x,t,z,dx,dy,dz [m]
Coords['FWDPR'] = [215, -25, 0, 20, 50, 19] #x,t,z,dx,dy,dz [m]
Coords['CCR'] = [10, -20, 37, 20, 5, 3] #x,t,z,dx,dy,dz [m]
Coords['Cabin'] = [10, -20, 45, 20, 40, 15] #x,t,z,dx,dy,dz [m]
Coords['RecreationRoom'] = [10, -20, 55, 20, 40, 5] #x,t,z,dx,dy,dz [m]
Coords['MessRoom'] = [10, -20, 40, 20, 40, 5] #x,t,z,dx,dy,dz [m]


#Rows: 1:14, 15:17, 18:30, 31:35

M = np.zeros((len(Groups),len(Areas)))


for c in Cases:
    shManning = iExl['POB - ' + c]
    #manning information in H6:AD40
    for g in Groups:
        iG1 = g[1]+5
        iG2 = g[2]+5
        iG = Groups.index(g)
        for a in Areas:
            iA = Areas.index(a)+8
            for ig in range(iG1,iG2+1):                
                m = shManning.cell(ig,iA).value
                if m != None:
                    M[iG,iA-8] += m

#ManningMap for 'density' of crews in rectangular grid over the FPSO
XX = np.linspace(99,224,50)
DX = XX[1]-XX[0]
YY = np.linspace(-27,27,21)
DY = YY[1]-YY[0]
AreaPatch = DX*DY
MM = np.zeros((len(Groups),len(XX),len(YY)))
for g in Groups:
    iG = Groups.index(g)
    for i in range(0,len(XX)):
        for j in range(0,len(YY)):
            x = XX[i]
            y = YY[j]
            #check to which area (x,y) belongs to
            #find the area that contains (x,y)
            for a in Areas:
                x1 = Coords[a][0]
                x2 = Coords[a][0] + Coords[a][3]
                y1 = Coords[a][1]
                y2 = Coords[a][1]+Coords[a][4]
                # z1 = Coords[a][2]
                # z2 = Coords[a][2]+Coords[a][4]
                if ((x >= x1) and (x < x2) and (y >= y1) and (y < y2)):
                    print (x, y, " in ", a)
                    break

            iA = Areas.index(a)
            Area = Coords[a][3]*Coords[a][4]
            MM[iG,i,j] = M[iG,iA]/Area*AreaPatch


Resolution = 2
# for g in Groups:
    # iG = Groups.index(g)

iG = 2
f = MM[iG,:,].transpose()
# levels = [0.001, 0.01, 0.1]

fig,ax = plt.subplots()

# img = plt.imread("RubyFPSODeckA.jpg")
# ax.imshow(img, extent=[-5.92, 251.95, -32.43, 50.19])

# cs1=ax.contourf(XX,YY,f,levels,colors=['yellow','magenta'],alpha=0.7)
cs1=ax.contourf(XX,YY,f)
# cs1.cmap.set_under('white')
# cs1.cmap.set_over('red')
# cs2=ax.contour(XX,YY,f,levels,colors=('k',),linewidths=(3,))
ax.clabel(cs2,fmt='%3d',colors='k',fontsize=14)
# ax.scatter(pvloc[:,0],pvloc[:,1],c='red')
ax.set_aspect('equal')
ax.xaxis.set_major_locator(plt.FixedLocator([120, 141, 168, 193]))
ax.xaxis.set_major_formatter(plt.FixedFormatter(['2/3','3/4','4/5','S05']))
ax.yaxis.set_major_locator(plt.FixedLocator([-27, -3.1, 3.1, 27]))
ax.yaxis.set_major_formatter(plt.FixedFormatter(['ER_S','Tray_S','Tray_P','ER_P']))

# ax.xaxis.set_major_locator(plt.FixedLocator(XFramesAlleys))
# ax.xaxis.set_major_formatter(plt.FixedFormatter(TickFrameAlleys))
ax.xaxis.grid(b=True, linestyle='--',linewidth=2)
ttl = Groups[iG]
plt.title(ttl)
# fig.savefig(fn+'v02.png')
plt.show()

